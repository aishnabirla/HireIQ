import pandas as pd
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


def safe_text(text, limit=100):
    text = str(text)[:limit]
    return text.encode('latin-1', 'replace').decode('latin-1')


def results_to_dataframe(results):
    return pd.DataFrame([{
        "Rank": i + 1,
        "Name": r['name'],
        "Email": r['email'],
        "Final Score (%)": r['final_score'],
        "Skill Match (%)": r['skill_score'],
        "Matched Skills": ", ".join(r['matched_skills']) if r['matched_skills'] else "None",
        "Missing Skills": ", ".join(r['missing_skills']) if r['missing_skills'] else "None",
        "Education": r['education'] or "Not found",
        "Experience": r['experience'] or "Not found"
    } for i, r in enumerate(results)])


def export_csv(results):
    df = results_to_dataframe(results)
    return df.to_csv(index=False)


def export_excel(results, jd_title):
    df = results_to_dataframe(results)
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    header_fill = PatternFill(
        start_color="1A73E8",
        end_color="1A73E8",
        fill_type="solid"
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
        size=11
    )
    alt_fill = PatternFill(
        start_color="E8F0FE",
        end_color="E8F0FE",
        fill_type="solid"
    )
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"HireIQ - Results for: {jd_title}"
    title_cell.font = Font(bold=True, size=14, color="1A73E8")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    ws.append([])

    headers = list(df.columns)
    ws.append(headers)
    header_row = ws.max_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    for row_idx, row in enumerate(df.itertuples(index=False), 1):
        ws.append(list(row))
        current_row = ws.max_row
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            if fill:
                cell.fill = fill
            cell.alignment = center
            cell.border = thin_border

    col_widths = [6, 20, 25, 14, 14, 35, 35, 30, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A4'
    wb.save(output)
    return output.getvalue()


def export_pdf(results, jd_title):
    pdf = FPDF()
    pdf.set_margins(15, 15, 15)
    pdf.add_page()

    pdf.set_font("Arial", "B", 18)
    pdf.cell(0, 12, "HireIQ - AI Resume Screener", ln=True, align="C")
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 10, f"Job: {safe_text(jd_title, 60)}", ln=True, align="C")
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 8, f"Total Candidates Evaluated: {len(results)}", ln=True, align="C")
    pdf.ln(6)

    pdf.set_fill_color(26, 115, 232)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(10, 8, "#", fill=True)
    pdf.cell(50, 8, "Name", fill=True)
    pdf.cell(30, 8, "Final Score", fill=True)
    pdf.cell(30, 8, "Skill Match", fill=True)
    pdf.cell(60, 8, "Email", fill=True, ln=True)

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 9)
    for i, r in enumerate(results):
        if i % 2 == 0:
            pdf.set_fill_color(232, 240, 254)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.cell(10, 7, str(i+1), fill=True)
        pdf.cell(50, 7, safe_text(r['name'], 25), fill=True)
        pdf.cell(30, 7, f"{r['final_score']}%", fill=True)
        pdf.cell(30, 7, f"{r['skill_score']}%", fill=True)
        pdf.cell(60, 7, safe_text(r['email'], 30), fill=True, ln=True)

    pdf.ln(8)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 8, "Detailed Breakdown", ln=True)
    pdf.ln(2)

    for i, r in enumerate(results):
        pdf.set_font("Arial", "B", 10)
        pdf.cell(
            0, 7,
            f"#{i+1} {safe_text(r['name'], 30)} - {r['final_score']}%",
            ln=True
        )
        pdf.set_font("Arial", "", 9)
        matched = safe_text(
            ", ".join(r['matched_skills']) if r['matched_skills'] else "None",
            120
        )
        missing = safe_text(
            ", ".join(r['missing_skills']) if r['missing_skills'] else "None",
            120
        )
        edu = safe_text(r['education'] or "Not found", 120)
        exp = safe_text(r['experience'] or "Not found", 120)
        pdf.cell(0, 6, f"Matched Skills: {matched}", ln=True)
        pdf.cell(0, 6, f"Missing Skills: {missing}", ln=True)
        pdf.cell(0, 6, f"Education: {edu}", ln=True)
        pdf.cell(0, 6, f"Experience: {exp}", ln=True)
        pdf.ln(4)

    return bytes(pdf.output())